import openpyxl
from openpyxl.styles import PatternFill

# Wczytaj plik Excel
wb = openpyxl.load_workbook('plik.xlsx')
sheet = wb.active  # Zakłada, że pracujemy na aktywnym arkuszu

# Zdefiniuj kolor czerwony
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Przejdź przez komórki w określonej kolumnie (np. kolumna A)
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):  # Zakładając, że teksty zaczynają się od 2 wiersza
    for cell in row:
        if 'python' in str(cell.value).lower():  # Sprawdzamy, czy 'python' jest w komórce
            cell.fill = red_fill  # Zaznacz komórkę na czerwono

# Zapisz zmodyfikowany plik
wb.save('plik_zmieniony.xlsx')
